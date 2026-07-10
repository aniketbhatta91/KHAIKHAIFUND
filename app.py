"""KhaiKhai Fund — shared team fund tracker.

Flask backend with SQLite. Tracks donations (money in) and expenses
(money out), and exposes a small JSON API consumed by the frontend.
"""

import os
import sqlite3
from datetime import date, datetime
from io import BytesIO

from flask import (
    Flask,
    g,
    jsonify,
    render_template,
    request,
    send_file,
    send_from_directory,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- Database selection ----------------------------------------------------- #
# In production set DATABASE_URL (e.g. a Render Postgres instance) and the app
# uses Postgres for durable data. With no DATABASE_URL it falls back to a local
# SQLite file (path overridable via KHAIKHAI_DB). Same code, both places.
DATABASE_URL = os.environ.get("DATABASE_URL")
USE_PG = bool(DATABASE_URL)
DB_PATH = os.environ.get("KHAIKHAI_DB", os.path.join(BASE_DIR, "fund.db"))

if USE_PG:
    import psycopg2 as _pg
    import psycopg2.extras as _pg_extras

app = Flask(__name__)


# --------------------------------------------------------------------------- #
# Database helpers
# --------------------------------------------------------------------------- #
def get_db():
    """Return a per-request database connection (Postgres or SQLite)."""
    if "db" not in g:
        if USE_PG:
            g.db = _pg.connect(DATABASE_URL)
        else:
            g.db = sqlite3.connect(DB_PATH)
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


def _cursor(db):
    """A cursor whose rows support name access (row["col"]) on both backends."""
    if USE_PG:
        return db.cursor(cursor_factory=_pg_extras.RealDictCursor)
    return db.cursor()


def _sql(query):
    """Translate the '?' placeholder style to Postgres '%s' when needed."""
    return query.replace("?", "%s") if USE_PG else query


@app.teardown_appcontext
def close_db(_exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# --------------------------------------------------------------------------- #
# CORS — allow the mobile app (and Expo web preview) to call this API
# --------------------------------------------------------------------------- #
@app.before_request
def _short_circuit_preflight():
    if request.method == "OPTIONS":
        return ("", 204)


@app.after_request
def _add_cors_headers(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, DELETE, OPTIONS"
    return resp


_SQLITE_SCHEMA = """
    CREATE TABLE IF NOT EXISTS transactions (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        type       TEXT NOT NULL CHECK (type IN ('donation', 'expense')),
        label      TEXT NOT NULL,
        person     TEXT NOT NULL,
        amount     REAL NOT NULL CHECK (amount > 0),
        txn_date   TEXT NOT NULL,
        note       TEXT DEFAULT '',
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    )
"""

_PG_SCHEMA = """
    CREATE TABLE IF NOT EXISTS transactions (
        id         SERIAL PRIMARY KEY,
        type       TEXT NOT NULL CHECK (type IN ('donation', 'expense')),
        label      TEXT NOT NULL,
        person     TEXT NOT NULL,
        amount     DOUBLE PRECISION NOT NULL CHECK (amount > 0),
        txn_date   TEXT NOT NULL,
        note       TEXT DEFAULT '',
        created_at TIMESTAMP NOT NULL DEFAULT now()
    )
"""


def init_db():
    """Create the transactions table if it does not exist (both backends)."""
    if USE_PG:
        conn = _pg.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute(_PG_SCHEMA)
        conn.commit()
        cur.close()
        conn.close()
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.execute(_SQLITE_SCHEMA)
        conn.commit()
        conn.close()


# --------------------------------------------------------------------------- #
# Serialization
# --------------------------------------------------------------------------- #
def row_to_dict(row):
    return {
        "id": row["id"],
        "type": row["type"],
        "label": row["label"],
        "person": row["person"],
        "amount": round(row["amount"], 2),
        "date": row["txn_date"],
        "note": row["note"] or "",
        "created_at": str(row["created_at"]),
    }


def compute_summary(db):
    cur = _cursor(db)
    cur.execute(
        """
        SELECT
            COALESCE(SUM(CASE WHEN type='donation' THEN amount END), 0) AS collected,
            COALESCE(SUM(CASE WHEN type='expense'  THEN amount END), 0) AS spent
        FROM transactions
        """
    )
    r = cur.fetchone()
    collected = round(r["collected"], 2)
    spent = round(r["spent"], 2)
    return {
        "collected": collected,
        "spent": spent,
        "balance": round(collected - spent, 2),
    }


# --------------------------------------------------------------------------- #
# Routes
# --------------------------------------------------------------------------- #
@app.route("/")
def index():
    return render_template("index.html")


# ---- PWA: manifest + service worker served from the root scope ---- #
@app.route("/manifest.webmanifest")
def manifest():
    return send_from_directory(
        app.static_folder, "manifest.webmanifest",
        mimetype="application/manifest+json",
    )


@app.route("/service-worker.js")
def service_worker():
    resp = send_from_directory(
        app.static_folder, "service-worker.js",
        mimetype="application/javascript",
    )
    # Allow the worker (served from /static path on disk) to control the whole site.
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


@app.route("/api/transactions", methods=["GET"])
def list_transactions():
    db = get_db()
    cur = _cursor(db)
    cur.execute("SELECT * FROM transactions ORDER BY txn_date DESC, id DESC")
    rows = cur.fetchall()
    return jsonify({
        "transactions": [row_to_dict(r) for r in rows],
        "summary": compute_summary(db),
    })


@app.route("/api/summary", methods=["GET"])
def summary():
    return jsonify(compute_summary(get_db()))


@app.route("/api/transactions", methods=["POST"])
def add_transaction():
    data = request.get_json(silent=True) or {}

    txn_type = (data.get("type") or "").strip().lower()
    label = (data.get("label") or "").strip()
    person = (data.get("person") or "").strip()
    note = (data.get("note") or "").strip()
    txn_date = (data.get("date") or "").strip() or date.today().isoformat()

    # Validation
    errors = []
    if txn_type not in ("donation", "expense"):
        errors.append("type must be 'donation' or 'expense'")
    if not label:
        errors.append("label (name or reason) is required")
    if not person:
        errors.append("person is required")
    try:
        amount = round(float(data.get("amount")), 2)
        if amount <= 0:
            errors.append("amount must be greater than 0")
    except (TypeError, ValueError):
        amount = None
        errors.append("amount must be a number")
    try:
        date.fromisoformat(txn_date)
    except ValueError:
        errors.append("date must be in YYYY-MM-DD format")

    if errors:
        return jsonify({"error": "; ".join(errors)}), 400

    db = get_db()
    cur = _cursor(db)
    params = (txn_type, label, person, amount, txn_date, note)
    insert = _sql(
        "INSERT INTO transactions (type, label, person, amount, txn_date, note) "
        "VALUES (?, ?, ?, ?, ?, ?)"
    )
    if USE_PG:
        cur.execute(insert + " RETURNING id", params)
        new_id = cur.fetchone()["id"]
    else:
        cur.execute(insert, params)
        new_id = cur.lastrowid
    db.commit()
    cur.execute(_sql("SELECT * FROM transactions WHERE id = ?"), (new_id,))
    row = cur.fetchone()
    return jsonify({
        "transaction": row_to_dict(row),
        "summary": compute_summary(db),
    }), 201


@app.route("/api/transactions/<int:txn_id>", methods=["DELETE"])
def delete_transaction(txn_id):
    db = get_db()
    cur = _cursor(db)
    cur.execute(_sql("DELETE FROM transactions WHERE id = ?"), (txn_id,))
    db.commit()
    if cur.rowcount == 0:
        return jsonify({"error": "transaction not found"}), 404
    return jsonify({"deleted": txn_id, "summary": compute_summary(db)})


@app.route("/api/export", methods=["GET"])
def export_xlsx():
    """Build an .xlsx of the full ledger with a summary block on top."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    cur = _cursor(db)
    cur.execute("SELECT * FROM transactions ORDER BY txn_date DESC, id DESC")
    rows = cur.fetchall()
    summary = compute_summary(db)

    # Styling palette
    accent = PatternFill("solid", fgColor="6C8CFF")
    head_fill = PatternFill("solid", fgColor="1F232C")
    band = PatternFill("solid", fgColor="F2F4FA")
    white_bold = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = u'₹#,##0.00;[Red]-₹#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Ledger"

    # --- Title ---
    ws["A1"] = "KhaiKhai Fund — Transaction History"
    ws["A1"].font = Font(size=16, bold=True, color="1F232C")
    ws.merge_cells("A1:E1")
    ws["A2"] = "Generated " + datetime.now().strftime("%d %b %Y, %H:%M")
    ws["A2"].font = Font(size=10, color="8B93A3")
    ws.merge_cells("A2:E2")

    # --- Summary block ---
    summary_rows = [
        ("Current Balance", summary["balance"]),
        ("Total Collected", summary["collected"]),
        ("Total Spent", summary["spent"]),
    ]
    r = 4
    for label, value in summary_rows:
        lc, vc = ws.cell(r, 1), ws.cell(r, 2)
        lc.value = label
        lc.font = Font(bold=True, color="1F232C")
        lc.fill = band
        vc.value = value
        vc.number_format = money_fmt
        vc.font = Font(bold=True)
        vc.fill = band
        lc.border = vc.border = border
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1

    # --- Ledger header ---
    header_row = r + 1  # blank spacer row between summary and table
    headers = ["Date", "Type", "Name/Item", "Note", "Amount"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(header_row, col, name)
        c.fill = head_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
    ws.cell(header_row, 5).alignment = Alignment(horizontal="right", vertical="center")

    # --- Ledger rows ---
    rownum = header_row + 1
    for t in rows:
        # Expenses are written as negative so the column sums to the balance.
        signed = t["amount"] if t["type"] == "donation" else -t["amount"]
        values = [
            t["txn_date"],
            "Donation" if t["type"] == "donation" else "Expense",
            t["label"],
            t["note"] or "",
            round(signed, 2),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(rownum, col, v)
            c.border = border
            if rownum % 2 == 0:
                c.fill = band
        ws.cell(rownum, 5).number_format = money_fmt
        rownum += 1

    # --- Column widths ---
    widths = [14, 12, 28, 36, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(header_row + 1, 1)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = "khaikhai-fund-" + date.today().isoformat() + ".xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Initialize the database at import time so it also works under a WSGI
# server like gunicorn (where the __main__ block below does NOT run).
init_db()


if __name__ == "__main__":
    # Local development server.
    # Bind to 0.0.0.0 so phones on the same Wi-Fi can reach the API.
    # Override with HOST/PORT env vars if needed.
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "5000"))
    app.run(host=host, port=port, debug=True)
